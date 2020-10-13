# Authors of the project:
# 1-MachnevEgor_https://vk.com/machnev_egor
# 2-SchalimovDmitriy_https://vk.com/astronaut_without_spaceship
# 3-ArsenyKarimov_https://vk.com/id222338543
# 4-MihailMarkov_https://vk.com/mixxxxail
# Contacts in email:
# 1-meb.official.com@gmail.com
# 2-dmitriy-shalimov@yandex.ru
# 3-arseny.karimov@gmail.com
# 4-mihailmarkov2004@gmail.com

# import module
import openpyxl

# databases settings
authors_comments = ["# Authors of the project:", "# 1-MachnevEgor_https://vk.com/machnev_egor",
                    "# 2-SchalimovDmitriy_https://vk.com/astronaut_without_spaceship",
                    "# 3-ArsenyKarimov_https://vk.com/id222338543", "# 4-MihailMarkov_https://vk.com/mixxxxail",
                    "# Contacts in email:", "# 1-meb.official.com@gmail.com", "# 2-dmitriy-shalimov@yandex.ru",
                    "# 3-arseny.karimov@gmail.com", "# 4-mihailmarkov2004@gmail.com"]
cell_separator = " | "


# standardizing the database if it is empty
def standardization_users_database(database_source):
    # opening the database
    data_memory = open(database_source, "r+", encoding="UTF-8")
    file_lines_array = list(data_memory.readlines())
    # sending data to the terminal
    print("--<Old data:" + "\n" + "".join(file_lines_array))
    # standardization
    if len(file_lines_array) < len(authors_comments):
        # sending data to the terminal
        print("-----------------------------------------------")
        print("!!! ERROR: Users database - BROKEN or empty !!!")
        print("Solution: Return to the initial form")
        print("-----------------------------------------------")
        # clear file data
        data_memory.truncate(0)
        data_memory.seek(0)
        # recording a layout
        data_memory.write("\n".join(authors_comments))
    # closing the database
    data_memory.close()


# transferring .xlsx to .txt users databse
def database_interpreter(excel_source, sheet_name, columns, source_for_user, columns_for_user, extra_cells, txt_source):
    standardization_users_database(database_source=txt_source)
    try:
        # open excel file
        excel_document = openpyxl.load_workbook(excel_source)
        # full name - import data from a graph and transfer it to a separate array
        full_name = []
        sheet = excel_document[sheet_name]
        for row in sheet[f"{columns[0]}1":f"{columns[0]}{sheet.max_row}"]:
            for cell in row:
                full_name.append(str(cell.value))
        # user id - import data from a graph and transfer it to a separate array
        user_id = []
        sheet = excel_document[sheet_name]
        for row in sheet[f"{columns[1]}1":f"{columns[1]}{sheet.max_row}"]:
            for cell in row:
                user_id.append(str(cell.value))
        # group name - import data from a graph and transfer it to a separate array
        group_name = []
        sheet = excel_document[sheet_name]
        for row in sheet[f"{columns[2]}1":f"{columns[2]}{sheet.max_row}"]:
            for cell in row:
                group_name.append(str(cell.value))
        # composing a string
        lines = []
        for i in range(len(full_name)):
            lines.append("\n" + str(full_name[i]) + cell_separator + str(user_id[i]) + cell_separator + str(
                source_for_user) + cell_separator + str(group_name[i]) + cell_separator + str(
                columns_for_user) + cell_separator + str(extra_cells))
        # opening the database
        data_memory = open(txt_source, "r+", encoding="UTF-8")
        file_lines_array = list(data_memory.readlines())
        # clear file data
        data_memory.truncate(0)
        data_memory.seek(0)
        # entering user data into the database and rewriting last data
        data_memory.write("".join(file_lines_array) + "".join(lines))
        # closing the database
        data_memory.close()
        # sending data to the terminal
        print("--<Done")
        print("--<All data:" + "".join(file_lines_array) + "".join(lines))
        print("--<New data:" + "".join(lines))
    except Exception as E:
        # sending data to the terminal
        print("!!! ERROR: .xlsx - not found or broken !!!")
        print(f"Reason: {E}")


# start work - make sure everything is correct
database_interpreter(excel_source="usersExcelDatabase/###xlsx", sheet_name="Лист1",
                     columns=["#", "#", "#"], source_for_user="##class",
                     columns_for_user="['A', 'B', 'E']", extra_cells=1, txt_source="UsersDatabase.txt")

# Authors of the project:
# 1-MachnevEgor_https://vk.com/machnev_egor
# 2-SchalimovDmitriy_https://vk.com/astronaut_without_spaceship
# 3-ArsenyKarimov_https://vk.com/id222338543
# 4-MihailMarkov_https://vk.com/mixxxxail
# Contacts in email:
# 1-meb.official.com@gmail.com
# 2-dmitriy-shalimov@yandex.ru
# 3-arseny.karimov@gmail.com
# 4-mihailmarkov2004@gmail.com
