# Authors of the project:
# 1-MachnevEgor_https://vk.com/machnev_egor
# 2-SchalimovDmitriy_https://vk.com/astronaut_without_spaceship
# 3-ArsenyKarimov_https://vk.com/id222338543
# 4-MihailMarkov_https://vk.com/mixxxxail
# Contacts in email:
# 1-meb.official.com@gmail.com
# 2-dmitriy-shalimov@yandex.ru
# 3-arseny.karimov@gmail.com
# 4-mihailmarkov2004@gmail.com

# import of the main module for further work with tables
import os as os
import openpyxl as openpyxl
import pickle as pickle


# the main function for analyzing the user's schedule
def selective_data_search(excel_source, columns, extra_cells, sheet_name, start_data, end_data,
                          excel_database_source="workWithExcelFile/excelDatabase",
                          unnecessary_words_or_sentences=["См.Таблицу После Субботы"], teacher_key="TEACHERS",
                          importance_of_the_error=True):
    # declaring a variable for further work with it
    user_schedule = ["Расписание на заданный день:"]
    # sending data to the terminal about the passed values, if required
    if importance_of_the_error != False:
        print(
            f"Schedule source: {excel_source}/{sheet_name}.xlsx({columns}, {extra_cells}, [{start_data}, {end_data}])")
    try:
        # opening a file for later work with it
        excel_document = openpyxl.load_workbook(f"{excel_database_source}/{excel_source}/{sheet_name}.xlsx")
        # days - import data from a graph and transfer it to a separate array
        days_data_array = []
        sheet = excel_document.get_sheet_by_name(sheet_name)
        for row in sheet[f"{columns[0]}1":f"{columns[0]}{sheet.max_row}"]:
            for cell in row:
                days_data_array.append(str(cell.value))
        # lessons - import data from a graph and transfer it to a separate array
        lessons_data_array = []
        sheet = excel_document.get_sheet_by_name(sheet_name)
        for row in sheet[f"{columns[1]}1":f"{columns[1]}{sheet.max_row}"]:
            for cell in row:
                lessons_data_array.append(str(cell.value))
        # teachers - import data from a graph and transfer it to a separate array
        teachers_data_array = []
        sheet = excel_document.get_sheet_by_name(sheet_name)
        for row in sheet[f"{columns[2]}1":f"{columns[2]}{sheet.max_row}"]:
            for cell in row:
                teachers_data_array.append(str(cell.value))
        # cabinets - import data from a graph and transfer it to a separate array
        cabinets_data_array = []
        sheet = excel_document.get_sheet_by_name(sheet_name)
        for row in sheet[f"{columns[3]}1":f"{columns[3]}{sheet.max_row}"]:
            for cell in row:
                cabinets_data_array.append(str(cell.value))
        # additional information - import data from a graph and transfer it to a separate array
        additional_information_data_array = []
        sheet = excel_document.get_sheet_by_name(sheet_name)
        for row in sheet[f"{columns[4]}1":f"{columns[4]}{sheet.max_row}"]:
            for cell in row:
                additional_information_data_array.append(str(cell.value))
        # search for relevant information
        lessons_output_data_array = []
        teachers_output_data_array = []
        cabinets_output_data_array = []
        additional_information_output_data_array = []
        for quantity_checks in range(len(days_data_array)):
            if days_data_array[quantity_checks].lower() == start_data.lower():
                for quantity_recording_data in range(len(lessons_data_array) - quantity_checks - 1 - extra_cells):
                    if lessons_data_array[
                        quantity_recording_data + quantity_checks + 1 + extra_cells].lower() == end_data.lower():
                        break
                    # lessons - writing the necessary information
                    lessons_output_data_array.append(
                        lessons_data_array[quantity_recording_data + quantity_checks + 1 + extra_cells].title())
                    # teachers - data analysis and writing the necessary information
                    if (teachers_data_array[
                            quantity_recording_data + quantity_checks + 1 + extra_cells] != "None") and (
                            teachers_data_array[
                                quantity_recording_data + quantity_checks + 1 + extra_cells].title() not in unnecessary_words_or_sentences):
                        teachers_output_data_array.append(
                            teachers_data_array[quantity_recording_data + quantity_checks + 1 + extra_cells].title())
                    else:
                        if excel_source != teacher_key:
                            teachers_output_data_array.append("Преподаватель не указан")
                        else:
                            teachers_output_data_array.append("Предмет не указан")
                    # cabinets - data analysis and writing the necessary information
                    if (cabinets_data_array[
                            quantity_recording_data + quantity_checks + 1 + extra_cells] != "None") and (
                            cabinets_data_array[
                                quantity_recording_data + quantity_checks + 1 + extra_cells].title() not in unnecessary_words_or_sentences):
                        try:
                            cabinets_output_data_array.append(int(float(cabinets_data_array[
                                                                            quantity_recording_data + quantity_checks + 1 + extra_cells])))
                        except Exception as E:
                            cabinets_output_data_array.append(cabinets_data_array[
                                                                  quantity_recording_data + quantity_checks + 1 + extra_cells].title())
                    else:
                        if excel_source != teacher_key:
                            cabinets_output_data_array.append("Узнавать у классного руководителя")
                        else:
                            cabinets_output_data_array.append("Номер кабинета узнавать у администрации")
                    # additional information - writing the necessary information
                    additional_information_output_data_array.append(additional_information_data_array[
                                                                        quantity_recording_data + quantity_checks + 1 + extra_cells].title())
        # building a schedule
        window_in_the_first_lesson = True
        for quantity_transfers in range(len(lessons_output_data_array)):
            if lessons_output_data_array[quantity_transfers].upper() != "ОКНО":
                if window_in_the_first_lesson == True:
                    window_in_the_first_lesson = False
                    if quantity_transfers + 1 != 2:
                        user_schedule.append(f"👉Начало занятий с {quantity_transfers + 1} урока👈")
                    else:
                        user_schedule.append(f"👉Начало занятий со {quantity_transfers + 1} урока👈")
                if additional_information_output_data_array[quantity_transfers] != "None":
                    user_schedule.append(
                        f"{quantity_transfers + 1}. 🧠{additional_information_output_data_array[quantity_transfers]} ({teachers_output_data_array[quantity_transfers]} & {cabinets_output_data_array[quantity_transfers]})")
                else:
                    user_schedule.append(
                        f"{quantity_transfers + 1}. {lessons_output_data_array[quantity_transfers]} ({teachers_output_data_array[quantity_transfers]} & {cabinets_output_data_array[quantity_transfers]})")
            elif window_in_the_first_lesson == False:
                user_schedule.append(f"{quantity_transfers + 1}. ОКНО (Можно отдохнуть в коворкинге)")
        # answer if there is nothing on this day
        if user_schedule == ["Расписание на заданный день:"]:
            if excel_source != teacher_key:
                user_schedule = "Кажись в этот день технопарк🙃"
            else:
                user_schedule = "В этот день нет занятий✨"
        else:
            user_schedule = "\n".join(user_schedule)
    except Exception as E:
        # sending error data to the terminal, if required
        if importance_of_the_error != False:
            print(f"!!! ERROR: Broken user data for excel searcher !!!")
            print(f"Reason: {E}")
        # generating an error message to notify the user
        user_schedule = "Очень странно - ты есть в базе, но некоторые данные неправильные. Напиши в основную беседу, прикреплённую к сообществу - там тебе помогут решить данную проблему😬\nhttps://vk.me/join/FhSVyJp7fYT0fM805_KTHNWPctDNa79JGsI="
    # returning the final result
    return user_schedule


# function for getting a dictionary with all groups and their users
def dictionary_of_groups_and_their_users(exception_words=["ФИО"], exception_folders=["GUESTS", "TEACHERS"],
                                         database_source="workWithExcelFile/excelDatabase",
                                         dump_source="workWithExcelFile/DictionaryDump.txt"):
    try:
        # check the old dump on the integrity of the dictionary, returning the dictionary if all is well
        return pickle.load(open(dump_source, "rb+"))
    # if the database is broken, reset the data, collect new ones
    except Exception as E:
        # sending data to the terminal about a broken dump
        print(f"!!! ERROR: Broken dump with groups and their users (scan and write new data) !!!")
        # the generation of paths to each file with a table
        groups_and_their_users = dict()
        for file_source in [f"{database_source}/{folder_name}/{file_name}" for folder_name in
                            os.listdir(database_source) for file_name in os.listdir(f"{database_source}/{folder_name}")
                            if folder_name not in exception_folders and "xlsx" in file_name.split(".")]:
            # opening a new table and then collecting data by column
            excel_document = openpyxl.load_workbook(file_source)
            sheet = excel_document.get_sheet_by_name(file_source.split("/")[-1][:-5])
            content_in_the_table = dict()
            for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                content_in_the_table.update(dict({letter: []}))
                for row in sheet[f"{letter}1":f"{letter}{sheet.max_row}"]:
                    for cell in row:
                        if cell.value != None and cell.value not in exception_words:
                            content_in_the_table[letter].append(cell.value)
            # search for the last column where the users that the group belongs to are stored
            last_content_column = "A"
            for letter_number in range(len(list(content_in_the_table.keys()))):
                if content_in_the_table[list(content_in_the_table.keys())[letter_number]] == [] and \
                        content_in_the_table[
                            list(content_in_the_table.keys())[letter_number - 1]] != []:
                    last_content_column = list(content_in_the_table.keys())[letter_number - 1]
            # assigning each group to its students, filling the dictionary
            groups_and_their_users.update(dict({file_source.split("/")[-1][:-5]: []}))
            for user in content_in_the_table[last_content_column]:
                groups_and_their_users[file_source.split("/")[-1][:-5]].append(user)
        # saving a new dictionary in the database and returning a new dictionary that was reassembled
        pickle.dump(groups_and_their_users, open(dump_source, "rb+"))
        return groups_and_their_users


# function for creating an array that stores all the groups in which the user is listed
def user_and_his_groups_groups(user_name, dump_with_groups=dictionary_of_groups_and_their_users()):
    return [group_name for group_name in dump_with_groups.keys() for group_users in dump_with_groups[group_name] if
            user_name in group_users]

# Authors of the project:
# 1-MachnevEgor_https://vk.com/machnev_egor
# 2-SchalimovDmitriy_https://vk.com/astronaut_without_spaceship
# 3-ArsenyKarimov_https://vk.com/id222338543
# 4-MihailMarkov_https://vk.com/mixxxxail
# Contacts in email:
# 1-meb.official.com@gmail.com
# 2-dmitriy-shalimov@yandex.ru
# 3-arseny.karimov@gmail.com
# 4-mihailmarkov2004@gmail.com
